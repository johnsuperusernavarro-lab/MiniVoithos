"""
reportes/generar_excel.py
Generador del reporte Excel de auditoría con siete hojas.

Código de colores:
  Rojo    (#C00000) — problemas críticos: facturas/retenciones no registradas
  Naranja (#C55A11) — advertencias: registros en sistema sin respaldo en SRI
  Verde   (#375623) — datos en orden: coincidencias correctas
  Azul    (#2E75B6) — encabezados generales
  Oscuro  (#1F3864) — título de portada RESUMEN
"""
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


# ── Helpers de estilo ──────────────────────────────────────────────────────────

def _relleno(color_hex):
    """Crea un PatternFill sólido con el color indicado en hexadecimal."""
    return PatternFill(fill_type='solid', fgColor=color_hex)


def _escribir_dataframe(ws, df, fila_inicio, nota=None, color_cabecera='2E75B6'):
    """
    Escribe un DataFrame en la hoja ws a partir de fila_inicio.
    Si se proporciona una nota, la escribe en gris itálico antes del encabezado.
    Devuelve el número de la siguiente fila disponible.
    """
    fila = fila_inicio

    if nota:
        celda = ws.cell(row=fila, column=1, value=nota)
        celda.font = Font(color='808080', italic=True, size=10)
        fila += 1

    if df is None or df.empty:
        celda = ws.cell(row=fila, column=1, value='Sin registros en esta categoría.')
        celda.font = Font(color='808080', size=10)
        return fila + 1

    # Encabezados
    for col_i, nombre_col in enumerate(df.columns, 1):
        celda = ws.cell(row=fila, column=col_i, value=nombre_col)
        celda.fill      = _relleno(color_cabecera)
        celda.font      = Font(bold=True, color='FFFFFF', size=10)
        celda.alignment = Alignment(horizontal='center')
    fila += 1

    # Datos
    for _, fila_datos in df.iterrows():
        for col_i, valor in enumerate(fila_datos, 1):
            celda      = ws.cell(row=fila, column=col_i, value=valor)
            celda.font = Font(size=10)
        fila += 1

    return fila


# ── Función principal ──────────────────────────────────────────────────────────

def generar_reporte(compras_result, retenciones_result, output_path):
    """
    Genera el archivo Excel de auditoría con las siguientes hojas:
      1. RESUMEN             — métricas clave de compras y retenciones
      2. COMPRAS - No en Sistema
      3. COMPRAS - Solo en Sistema
      4. COMPRAS - Coincidencias
      5. RET - No en Sistema
      6. RET - Sin XML
      7. RET - Coincidencias

    Args:
        compras_result (dict):     Resultado de comparar_compras().
        retenciones_result (dict): Resultado de comparar_retenciones().
        output_path (str):         Ruta completa del archivo .xlsx a generar.

    Returns:
        str: La misma output_path si todo fue exitoso.
    """
    wb = openpyxl.Workbook()
    stats_c = compras_result.get('stats', {})
    stats_r = retenciones_result.get('stats', {})

    # ── Hoja RESUMEN ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'RESUMEN'

    # Título principal
    celda_titulo = ws.cell(row=1, column=1, value='VOITHOS — Auditoría Contable SRI')
    celda_titulo.font = Font(bold=True, color='FFFFFF', size=14)
    celda_titulo.fill = _relleno('1F3864')
    ws.merge_cells('A1:D1')

    # Fecha de generación
    ws.cell(row=2, column=1,
            value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = \
        Font(color='808080', italic=True, size=10)

    fila = 4

    def _seccion(titulo):
        nonlocal fila
        c = ws.cell(row=fila, column=1, value=titulo)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = _relleno('2E75B6')
        fila += 1

    def _linea(etiqueta, valor, alerta=False):
        nonlocal fila
        ws.cell(row=fila, column=1, value=etiqueta).font = Font(size=10)
        celda_val = ws.cell(row=fila, column=2, value=valor)
        es_alerta = alerta and isinstance(valor, int) and valor > 0
        celda_val.font = Font(bold=True, size=10,
                              color='C00000' if es_alerta else '000000')
        fila += 1

    _seccion('COMPRAS')
    _linea('Facturas en XMLs descargados',          stats_c.get('xml',             0))
    _linea('Facturas en Sistema',                   stats_c.get('sistema',         0))
    _linea('Facturas en TXT del SRI',               stats_c.get('txt',             0))
    _linea('Coinciden en los 3 orígenes',           stats_c.get('en_todos',        0))
    _linea('En SRI pero NO en Sistema',             stats_c.get('no_en_sistema',   0), alerta=True)
    _linea('Solo en Sistema (sin SRI / sin XML)',   stats_c.get('solo_en_sistema', 0), alerta=True)

    fila += 1
    _seccion('RETENCIONES')
    _linea('Retenciones en XMLs',                  stats_r.get('ret_xml',    0))
    _linea('Ventas con retención en Sistema',       stats_r.get('vp_con_ret', 0))
    _linea('Coinciden',                            stats_r.get('en_ambos',   0))
    _linea('XML sin registro en Sistema',          stats_r.get('sin_sistema', 0), alerta=True)
    _linea('Sistema sin XML de respaldo',          stats_r.get('sin_xml',    0), alerta=True)

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 12

    # ── Hojas de detalle ──────────────────────────────────────────────────────
    hojas = [
        ('COMPRAS - No en Sistema',
         compras_result.get('no_en_sistema'),
         'Facturas en el TXT del SRI que NO están registradas en el Sistema',
         'C00000'),
        ('COMPRAS - Solo en Sistema',
         compras_result.get('solo_en_sistema'),
         'Facturas en el Sistema que NO aparecen en el TXT del SRI ni en los XMLs',
         'C55A11'),
        ('COMPRAS - Coincidencias',
         compras_result.get('coincidencias'),
         'La columna "Diferencia" muestra Total XML − Total Sistema. Cero = sin discrepancia.',
         '375623'),
        ('RET - No en Sistema',
         retenciones_result.get('sin_sistema'),
         'Retenciones en XML que NO tienen registro en Ventas_Personalizado',
         'C00000'),
        ('RET - Sin XML',
         retenciones_result.get('sin_xml'),
         'Ventas con retención registrada en Sistema pero SIN XML de respaldo',
         'C55A11'),
        ('RET - Coincidencias',
         retenciones_result.get('coincidencias'),
         '"Dif. IVA" y "Dif. IR" = valor XML − valor Sistema. Cero = sin discrepancia.',
         '375623'),
    ]

    for titulo, df, nota, color in hojas:
        ws_hoja = wb.create_sheet(titulo)
        _escribir_dataframe(ws_hoja, df, 3, nota=nota, color_cabecera=color)

    wb.save(output_path)
    print(f'\n   ✓  Reporte guardado: {output_path}')
    return output_path
