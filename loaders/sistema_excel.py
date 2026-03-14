"""
loaders/sistema_excel.py
Cargador del ReporteComprasVentas.xlsx exportado desde el sistema contable.

El archivo tiene una sola hoja activa con dos secciones separadas por marcadores:
  - La celda que contiene 'COMPRAS' marca el inicio de la sección de compras.
  - La celda que contiene 'VENTAS' marca el inicio de la sección de ventas.

Layout real (ejemplo enero 2026):
  Fila 0:   Razón Social del contribuyente
  Fila 1:   Título del reporte
  Fila 2:   Período
  Fila 5:   Marcador "COMPRAS"
  Fila 6:   Cabeceras de columna
  Filas 7+: Datos de compras
  Fila N:   Marcador "VENTAS"
  Fila N+1: Cabeceras de columna de ventas
  Filas +2: Datos de ventas
"""
import os
import re

import openpyxl
import pandas as pd

from util.normalizacion import formatear_num_documento
from util.validacion import fila_parece_cabecera


def cargar_sistema(path):
    """
    Lee el ReporteComprasVentas.xlsx y devuelve (df_compras, df_ventas).

    Detecta automáticamente las secciones COMPRAS y VENTAS buscando esas
    palabras en cualquier celda de la hoja activa.

    Defensas implementadas:
    - Detecta y reporta archivos .xls (formato antiguo) con mensaje claro.
    - Detecta archivos protegidos con contraseña.
    - Maneja workbooks sin hoja activa definida.
    - Reporta si no se detectan marcadores COMPRAS/VENTAS.

    Args:
        path (str): Ruta al archivo Excel del sistema contable.

    Returns:
        tuple: (df_compras, df_ventas) — ambos pd.DataFrame normalizados.
               df_ventas puede ser vacío si no se detecta la sección VENTAS.

    Raises:
        ValueError: Con mensaje amigable para el usuario en caso de error.
    """
    nombre = os.path.basename(path)

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        msg_lower = str(e).lower()
        if 'not a zip' in msg_lower or 'bad magic' in msg_lower:
            raise ValueError(
                f"'{nombre}' no es un archivo Excel válido (.xlsx).\n"
                "Si el archivo tiene extensión .xls (formato antiguo), "
                "ábrelo en Excel → Guardar como → Libro de Excel (.xlsx)."
            ) from e
        if 'encrypted' in msg_lower or 'password' in msg_lower:
            raise ValueError(
                f"'{nombre}' está protegido con contraseña.\n"
                "En Excel: Revisar → Proteger libro → Quitar contraseña, y vuelve a intentar."
            ) from e
        raise ValueError(f"No se pudo abrir '{nombre}': {e}") from e

    # ── Hoja activa ───────────────────────────────────────────────────────────
    ws = wb.active
    if ws is None:
        if wb.sheetnames:
            ws = wb[wb.sheetnames[0]]
            print(f"   ⚠  Hoja activa no definida — usando primera hoja: '{ws.title}'")
        else:
            raise ValueError(f"'{nombre}' no contiene hojas de datos.")

    filas = list(ws.values)
    if not filas:
        raise ValueError(f"'{nombre}': la hoja '{ws.title}' está vacía.")

    # ── Detectar marcadores de sección ────────────────────────────────────────
    compras_inicio = ventas_inicio = None
    for i, fila in enumerate(filas):
        # Solo considerar celdas cortas (≤ 30 chars) como marcadores de sección,
        # para evitar falsos positivos en celdas de Detalle que contengan las palabras.
        marcadores = [str(c).strip().upper() for c in fila
                      if c is not None and len(str(c).strip()) <= 30]
        if compras_inicio is None and any(m == 'COMPRAS' or m.startswith('COMPRAS') for m in marcadores):
            compras_inicio = i
        # Solo tomar el PRIMER marcador VENTAS después de COMPRAS para evitar
        # que descripciones de datos con la palabra "VENTAS" desplacen el índice
        if ventas_inicio is None and any(m == 'VENTAS' or m.startswith('VENTAS') for m in marcadores):
            ventas_inicio = i

    if compras_inicio is None:
        print(f"   ⚠  '{nombre}': no se detectó marcador COMPRAS — "
              "se intentará leer todo el contenido como compras")

    if compras_inicio is not None and ventas_inicio is not None and ventas_inicio > compras_inicio:
        df_compras = _seccion_a_df(filas[compras_inicio:ventas_inicio], nombre)
        df_ventas  = _seccion_a_df(filas[ventas_inicio:], nombre)
    elif compras_inicio is not None:
        df_compras = _seccion_a_df(filas[compras_inicio:], nombre)
        df_ventas  = pd.DataFrame()
    else:
        # Sin marcadores: intentar leer todo como compras
        df_compras = _seccion_a_df(filas, nombre)
        df_ventas  = pd.DataFrame()

    print(f"   ✓  {len(df_compras)} compras / {len(df_ventas)} ventas en sistema")
    return df_compras, df_ventas


def _seccion_a_df(filas, nombre_archivo=""):
    """
    Convierte una lista de filas en un DataFrame, detectando la fila de
    cabeceras de forma robusta:

    Estrategia 1 (preferida): busca la primera fila con ≥ 2 columnas
                              que coincidan con cabeceras contables conocidas.
    Estrategia 2 (fallback):  busca la primera fila cuyas celdas empiecen
                              con letras mayúsculas (comportamiento original).

    Normaliza las columnas NO_COMPROBANTE y NO_AUTORIZACION.
    """
    fila_cabecera = None

    # Estrategia 1: coincidencia con columnas contables conocidas
    for i, fila in enumerate(filas):
        if fila_parece_cabecera(fila):
            fila_cabecera = i
            break

    # Estrategia 2 (fallback): celda que empieza con mayúsculas
    if fila_cabecera is None:
        for i, fila in enumerate(filas):
            if any(isinstance(c, str) and re.match(r'^[A-Z]{2,}', str(c))
                   for c in fila if c):
                fila_cabecera = i
                break

    if fila_cabecera is None or fila_cabecera + 1 >= len(filas):
        if nombre_archivo:
            print(f"   ⚠  '{nombre_archivo}': no se detectó fila de cabeceras en una sección")
        return pd.DataFrame(columns=['SERIE', 'CLAVE'])

    cabeceras  = [
        str(c) if c is not None else f'COL_{i}'
        for i, c in enumerate(filas[fila_cabecera])
    ]
    filas_data = [f for f in filas[fila_cabecera + 1:] if any(c is not None for c in f)]

    if not filas_data:
        return pd.DataFrame(columns=cabeceras + ['SERIE', 'CLAVE'])

    # Alinear longitud de filas con cabeceras (celdas fusionadas pueden causar desajuste)
    n_cols = len(cabeceras)
    filas_norm = []
    for fila in filas_data:
        fila_lista = list(fila)
        if len(fila_lista) < n_cols:
            fila_lista += [None] * (n_cols - len(fila_lista))
        filas_norm.append(fila_lista[:n_cols])

    df = pd.DataFrame(filas_norm, columns=cabeceras)

    # Buscar columnas clave por posición (no por nombre) para tolerar duplicados.
    # El Excel del sistema contable puede tener "No. Autorización" dos veces
    # (una para la factura, otra para la retención) y otros nombres duplicados.
    # Mapa: nombre_normalizado → fragmentos a buscar en el nombre de columna (uppercase)
    _BUSCAR = {
        'RUC_EMISOR':      ['RUC'],
        'RAZON_SOCIAL':    ['RAZON SOCIAL', 'RAZÓN SOCIAL', 'RAZON_SOCIAL'],
        'FECHA_EMISION':   ['F. EMISION', 'F. EMISIÓN', 'FECHA EMISION', 'FECHA EMISIÓN'],
        'IVA_GASTO':       ['IVA GASTO'],
        'SUBTOTAL':        ['SUBTOTAL'],
        'TOTAL':           ['TOTAL'],
        'IVA':             ['IVA'],
    }
    # Para cada columna destino, asignar la PRIMERA columna fuente que coincida
    for destino, variantes in _BUSCAR.items():
        if destino in df.columns:
            continue  # ya existe
        for idx, col in enumerate(cabeceras):
            col_up = str(col).strip().upper()
            if any(v in col_up for v in variantes):
                df[destino] = df.iloc[:, idx]
                break

    # SERIE: extraer de "No. Comprobante" (primera ocurrencia)
    if 'SERIE' not in df.columns:
        for idx, col in enumerate(cabeceras):
            col_up = str(col).strip().upper()
            if 'COMPROBANTE' in col_up or 'NO. COMPROBANTE' in col_up:
                df['SERIE'] = (
                    df.iloc[:, idx].astype(str).str.strip()
                    .apply(formatear_num_documento)
                )
                break

    # CLAVE_ACCESO: extraer de la PRIMERA "No. Autorización" (la de la factura)
    # Se guarda como CLAVE_ACCESO para no interferir con la columna CLAVE
    # que comparar_compras construye como RUC_EMISOR|SERIE
    if 'CLAVE_ACCESO' not in df.columns:
        for idx, col in enumerate(cabeceras):
            col_up = str(col).strip().upper()
            if 'AUTORIZAC' in col_up:
                df['CLAVE_ACCESO'] = df.iloc[:, idx].astype(str).str.strip()
                break

    # Convertir columnas numéricas donde sea posible
    # Iterar por índice para evitar el problema de nombres duplicados
    # (columnas duplicadas harían que df[col] devuelva un DataFrame, no una Series)
    for idx in range(len(df.columns)):
        serie = df.iloc[:, idx]
        if serie.dtype != object:
            continue
        convertido = pd.to_numeric(serie, errors='coerce')
        if convertido.notna().sum() >= len(df) * 0.5:
            # fillna(0) mantiene tipo uniforme float; restaurar strings crearía
            # columnas mixtas (float + str) que rompen operaciones aritméticas posteriores
            df.iloc[:, idx] = convertido.fillna(0)

    return df
